"""
Test suite for Excel File Persistence Manager

Tests the immediate sync functionality to ensure the reported sync delay issue is resolved.
Specifically designed for Excel file operations in MAGK Excel automation.
"""
import pytest
import tempfile
import time
import threading
from pathlib import Path
from unittest.mock import Mock
import openpyxl

import sys
sys.path.append(str(Path(__file__).parent.parent / "src"))

from workflows.file_persistence_manager import ExcelFilePersistenceManager, ExcelFileRecord


class TestExcelFilePersistenceManager:
    """Test cases for the Excel file persistence manager with immediate sync."""
    
    def setup_method(self):
        """Set up test environment before each test."""
        self.temp_dir = tempfile.mkdtemp()
        self.excel_dir = Path(self.temp_dir) / "test_excel_storage"
        self.manager = ExcelFilePersistenceManager(
            excel_directory=str(self.excel_dir),
            auto_sync_interval=0.05  # Fast sync for testing
        )
        
        # Create a test Excel file
        self.test_excel_file = Path(self.temp_dir) / "test_workbook.xlsx"
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet['A1'] = "Test data for immediate sync verification"
        worksheet['B1'] = 42
        workbook.save(self.test_excel_file)
        workbook.close()
    
    def teardown_method(self):
        """Clean up after each test."""
        self.manager.shutdown()
        
        # Clean up temp directory
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)
    
    def test_immediate_sync_on_excel_open(self):
        """Test that Excel file opening syncs immediately without delay."""
        # Set up sync callback to track sync events
        sync_events = []
        
        def sync_callback(file_id, excel_record):
            sync_events.append((file_id, excel_record, time.time()))
        
        self.manager.add_sync_callback(sync_callback)
        
        # Record open time
        open_start = time.time()
        
        # Open Excel file
        file_id = self.manager.open_excel_file(str(self.test_excel_file))
        
        open_end = time.time()
        
        # Verify immediate sync
        assert file_id is not None
        assert len(sync_events) > 0, "Sync callback should be called immediately"
        
        # Check sync happened during open operation
        sync_time = sync_events[0][2]
        assert open_start <= sync_time <= open_end + 0.1, "Sync should happen immediately during open"
        
        # Verify Excel file info is immediately available
        excel_info = self.manager.get_excel_info(file_id)
        assert excel_info is not None
        assert excel_info.sync_status == "synced"
        assert excel_info.filename == "test_workbook.xlsx"
        assert excel_info.sheet_count > 0
        assert excel_info.is_open == True
    
    def test_excel_save_immediate_sync(self):
        """Test that Excel file saving syncs immediately."""
        # Open Excel file first
        file_id = self.manager.open_excel_file(str(self.test_excel_file))
        
        # Load workbook for editing
        workbook = openpyxl.load_workbook(self.test_excel_file)
        worksheet = workbook.active
        worksheet['C1'] = "Modified for sync test"
        
        # Set up sync callback
        sync_events = []
        def sync_callback(file_id, excel_record):
            sync_events.append((file_id, excel_record, time.time()))
        self.manager.add_sync_callback(sync_callback)
        
        # Save with immediate sync
        save_start = time.time()
        success = self.manager.save_excel_file(file_id, workbook)
        save_end = time.time()
        
        workbook.close()
        
        assert success, "Excel file save should succeed"
        assert len(sync_events) > 0, "Save should trigger sync callback"
        
        # Verify sync happened immediately
        sync_time = sync_events[0][2]
        assert save_start <= sync_time <= save_end + 0.1, "Save sync should be immediate"
        
        # Verify file status
        excel_info = self.manager.get_excel_info(file_id)
        assert excel_info.sync_status == "synced"
    
    def test_excel_sync_status_tracking(self):
        """Test that Excel sync status is properly tracked and reported."""
        # Open an Excel file
        file_id = self.manager.open_excel_file(str(self.test_excel_file))
        
        # Check sync status immediately
        sync_status = self.manager.get_sync_status()
        assert file_id in sync_status
        assert sync_status[file_id] == "synced"
        
        # Verify individual Excel file status
        excel_info = self.manager.get_excel_info(file_id)
        assert excel_info.sync_status == "synced"
        assert excel_info.is_open == True
    
    def test_multiple_excel_files_immediate_sync(self):
        """Test immediate sync with multiple Excel files opened quickly."""
        sync_events = []
        
        def sync_callback(file_id, excel_record):
            sync_events.append((file_id, excel_record.filename, time.time()))
        
        self.manager.add_sync_callback(sync_callback)
        
        # Create multiple test Excel files
        test_excel_files = []
        for i in range(3):
            excel_file = Path(self.temp_dir) / f"test_workbook_{i}.xlsx"
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet['A1'] = f"Test data for file {i}"
            workbook.save(excel_file)
            workbook.close()
            test_excel_files.append(excel_file)
        
        # Open Excel files rapidly
        open_start = time.time()
        file_ids = []
        
        for i, excel_file in enumerate(test_excel_files):
            file_id = self.manager.open_excel_file(str(excel_file))
            file_ids.append(file_id)
        
        open_end = time.time()
        
        # Allow small buffer for async callbacks
        time.sleep(0.1)
        
        # Verify all Excel files synced immediately
        assert len(sync_events) == 3, "All Excel files should trigger sync callbacks"
        assert len(file_ids) == 3, "All Excel files should be opened"
        
        # Check all sync events happened during open timeframe
        for file_id, filename, sync_time in sync_events:
            assert open_start <= sync_time <= open_end + 0.2, f"Excel file {filename} should sync immediately"
        
        # Verify all Excel files are in synced state
        sync_status = self.manager.get_sync_status()
        for file_id in file_ids:
            assert sync_status[file_id] == "synced"
    
    def test_excel_close_immediate_sync(self):
        """Test that Excel file closing syncs immediately."""
        # Open an Excel file first
        file_id = self.manager.open_excel_file(str(self.test_excel_file))
        assert self.manager.get_excel_info(file_id) is not None
        assert self.manager.get_excel_info(file_id).is_open == True
        
        # Close the Excel file
        close_start = time.time()
        success = self.manager.close_excel_file(file_id)
        close_end = time.time()
        
        assert success, "Excel file close should succeed"
        
        # Verify Excel file is marked as closed
        excel_info = self.manager.get_excel_info(file_id)
        assert excel_info is not None, "Excel file should still be tracked"
        assert excel_info.is_open == False, "Excel file should be marked as closed"
        assert excel_info.sync_status == "synced", "Closed Excel file should remain synced"
    
    def test_excel_force_sync_functionality(self):
        """Test that Excel force sync works immediately."""
        # Open an Excel file
        file_id = self.manager.open_excel_file(str(self.test_excel_file))
        
        # Force sync
        sync_start = time.time()
        success = self.manager.force_sync()
        sync_end = time.time()
        
        assert success, "Excel force sync should succeed"
        assert sync_end - sync_start < 0.5, "Excel force sync should complete quickly"
        
        # Verify Excel file still synced
        excel_info = self.manager.get_excel_info(file_id)
        assert excel_info.sync_status == "synced"
    
    def test_concurrent_excel_open_sync_safety(self):
        """Test that concurrent Excel file operations maintain sync integrity."""
        sync_events = []
        sync_lock = threading.Lock()
        
        def sync_callback(file_id, excel_record):
            with sync_lock:
                sync_events.append((file_id, excel_record.filename))
        
        self.manager.add_sync_callback(sync_callback)
        
        # Create multiple test Excel files
        test_excel_files = []
        for i in range(3):
            excel_file = Path(self.temp_dir) / f"concurrent_excel_{i}.xlsx"
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet['A1'] = f"Concurrent Excel content {i}"
            workbook.save(excel_file)
            workbook.close()
            test_excel_files.append(excel_file)
        
        # Open Excel files concurrently
        threads = []
        results = [None] * len(test_excel_files)
        
        def open_excel_file(index, file_path):
            try:
                file_id = self.manager.open_excel_file(str(file_path))
                results[index] = file_id
            except Exception as e:
                results[index] = f"Error: {e}"
        
        # Start concurrent opens
        for i, excel_file in enumerate(test_excel_files):
            thread = threading.Thread(target=open_excel_file, args=(i, excel_file))
            threads.append(thread)
            thread.start()
        
        # Wait for all opens to complete
        for thread in threads:
            thread.join(timeout=5.0)
        
        # Allow time for all sync callbacks
        time.sleep(0.2)
        
        # Verify all opens succeeded
        for i, result in enumerate(results):
            assert isinstance(result, str) and not result.startswith("Error"), f"Excel open {i} should succeed: {result}"
        
        # Verify all Excel files are synced
        sync_status = self.manager.get_sync_status()
        for file_id in results:
            if file_id and not file_id.startswith("Error"):
                assert sync_status.get(file_id) == "synced", f"Excel file {file_id} should be synced"
        
        # Verify sync events were recorded
        assert len(sync_events) == 3, "All Excel opens should trigger sync events"
    
    def test_excel_metadata_persistence_immediate(self):
        """Test that Excel metadata is persisted immediately after operations."""
        # Open an Excel file
        file_id = self.manager.open_excel_file(str(self.test_excel_file))
        
        # Check that metadata file exists immediately
        metadata_file = self.excel_dir / "excel_metadata.json"
        assert metadata_file.exists(), "Excel metadata file should exist immediately after open"
        
        # Create new manager instance to test persistence
        new_manager = ExcelFilePersistenceManager(excel_directory=str(self.excel_dir))
        
        try:
            # Verify Excel file is loaded from metadata
            excel_info = new_manager.get_excel_info(file_id)
            assert excel_info is not None, "Excel file should be loaded from persistent metadata"
            assert excel_info.filename == "test_workbook.xlsx"
            assert excel_info.sync_status == "synced"
            assert excel_info.sheet_count > 0
        finally:
            new_manager.shutdown()
    
    def test_excel_no_sync_delay_regression(self):
        """Regression test to ensure Excel sync delays don't reoccur."""
        # This test specifically addresses the original issue for Excel files
        
        sync_times = []
        
        def sync_callback(file_id, excel_record):
            sync_times.append(time.time())
        
        self.manager.add_sync_callback(sync_callback)
        
        # Open Excel file and measure sync timing
        open_start = time.time()
        file_id = self.manager.open_excel_file(str(self.test_excel_file))
        open_end = time.time()
        
        # Wait a short time for any async operations
        time.sleep(0.05)
        
        # Verify sync happened immediately
        assert len(sync_times) > 0, "Excel sync should happen immediately"
        sync_time = sync_times[0]
        
        # Sync should happen within open timeframe + small buffer
        sync_delay = sync_time - open_start
        assert sync_delay < 0.1, f"Excel sync delay ({sync_delay:.3f}s) should be minimal (< 0.1s)"
        
        # Excel file should be immediately available and synced
        excel_info = self.manager.get_excel_info(file_id)
        assert excel_info is not None
        assert excel_info.sync_status == "synced"
        assert excel_info.is_open == True
        
        print(f"✓ Excel sync completed in {sync_delay:.3f} seconds - Issue resolved!")


if __name__ == "__main__":
    pytest.main([__file__, "-v"])